"""Weekly timesheets.

Why this exists: when an employee misses a day, HR fills in the check-in and
check-out — but HR usually can't say what the person actually *did*. That
leaves days with hours and no work notes. The weekly timesheet is where the
employee completes those details, and where the manager signs them off.

Flow:

    employee completes the week (in-time, out-time and notes on every
    working day)                      ->  PENDING
    reporting manager approves        ->  APPROVED, and the whole week is
                                          written back into attendance —
                                          missing days are created
    reporting manager rejects         ->  REJECTED, employee fixes and
                                          resubmits

Two rules worth stating plainly, because they're what makes this trustworthy:

* Only the employee's REPORTING MANAGER may decide. HR sees everything
  org-wide but cannot approve — HR is the party that filled the gaps in the
  first place, and an approver signing off their own entries isn't approval.
* Hours are never taken from the client. They're recomputed from the
  approved in/out times with `classify_on_checkout`, the same function the
  live checkout path uses, so a timesheet day and a normally-worked day are
  classified identically.
"""

import io

from fastapi import (
    APIRouter, Depends, HTTPException, Query, UploadFile, File,
)
from fastapi.responses import StreamingResponse

from bson import ObjectId
from bson.errors import InvalidId

from datetime import datetime, timedelta, time
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from database import db
from utils.dependencies import (
    get_current_user,
    require_hr,
    require_manager_or_hr,
)
from utils.audit import log_audit
from utils.notify import notify_user  # push + in-app, both best-effort
from utils.ist import (
    now_ist_naive,
    today_ist_str,
    iso_naive,
    parse_wallclock_to_ist_naive,
)
from utils.attendance_rules import classify_on_checkout, is_weekend
from models.timesheet import (
    TimesheetSubmit,
    TimesheetDecision,
)


user_router = APIRouter()      # /timesheets/...
manager_router = APIRouter()   # /manager/timesheets/...
hr_router = APIRouter()        # /hr/timesheets/...


# Day types that carry no work, so they need no in/out time and no notes.
NON_WORKING_TYPES = {"LEAVE", "HOLIDAY", "WEEKLY_OFF"}


def _parse_date(s: str, field: str) -> datetime:
    try:
        return datetime.strptime(s, "%Y-%m-%d")
    except (TypeError, ValueError):
        raise HTTPException(400, f"Invalid {field} (use YYYY-MM-DD)")


def _week_dates(week_start: str) -> list[str]:
    start = _parse_date(week_start, "weekStart")
    if start.weekday() != 0:
        raise HTTPException(400, "weekStart must be a Monday")
    return [
        (start + timedelta(days=i)).strftime("%Y-%m-%d")
        for i in range(7)
    ]


async def _reporting_manager_id(user_id: str) -> Optional[str]:
    try:
        u = await db.users.find_one(
            {"_id": ObjectId(user_id)}, {"reportingManagerId": 1},
        )
    except (InvalidId, TypeError):
        return None
    return (u or {}).get("reportingManagerId")


async def _user_brief(user_id: Optional[str]) -> Optional[dict]:
    if not user_id:
        return None
    try:
        u = await db.users.find_one(
            {"_id": ObjectId(user_id)},
            {
                "name": 1,
                "email": 1,
                "employeeCode": 1,
                "reportingManagerId": 1,
            },
        )
    except (InvalidId, TypeError):
        return None
    if not u:
        return None
    return {
        "id": str(u["_id"]),
        "name": u.get("name"),
        "email": u.get("email"),
        "employeeCode": u.get("employeeCode"),
        "reportingManagerId": u.get("reportingManagerId"),
    }


async def _exempt_dates(user_id: str, dates: list[str]) -> dict[str, str]:
    """Dates in this week that don't require in/out/notes, with the reason.

    Without this, "every field on every day" would demand a check-in time on
    a Sunday and no week could ever be submitted.

    But a weekend/holiday/leave day the employee ACTUALLY WORKED is a working
    day. Exempting it on calendar grounds alone would hide real work: no notes
    requested, and the client collapses exempt rows entirely. So anything with
    attendance showing work is removed from this map at the end.
    """
    out: dict[str, str] = {}

    for d in dates:
        if is_weekend(d):
            out[d] = "Weekly off"

    async for h in db.holidays.find({"date": {"$in": dates}}):
        d = h.get("date")
        if d:
            out[d] = h.get("name") or "Holiday"

    # Approved leave overlapping the week. YYYY-MM-DD sorts chronologically,
    # so a plain string range comparison is correct here.
    async for lv in db.leave_requests.find(
        {
            "userId": user_id,
            "status": "APPROVED",
            "fromDate": {"$lte": dates[-1]},
            "toDate": {"$gte": dates[0]},
        }
    ):
        frm = lv.get("fromDate")
        to = lv.get("toDate") or frm
        if not frm:
            continue
        for d in dates:
            if frm <= d <= to:
                out[d] = "On leave"

    # Worked it? Then it's a working day, whatever the calendar says.
    # Someone who comes in on a Saturday still has to describe what they did,
    # and those hours still count.
    if out:
        async for r in db.attendance.find(
            {"userId": user_id, "date": {"$in": list(out.keys())}},
            {"date": 1, "checkIn": 1, "hoursWorked": 1, "attendanceType": 1},
        ):
            worked = bool(r.get("checkIn")) or float(
                r.get("hoursWorked", 0) or 0
            ) > 0
            if worked and (r.get("attendanceType") or "").upper() != "LEAVE":
                out.pop(r.get("date"), None)

    # Days that haven't happened yet. Applied LAST so nothing above can
    # un-exempt them: there is no work to describe for tomorrow, and
    # demanding it is what let a whole future week be fabricated and
    # approved into attendance.
    today = today_ist_str()
    for d in dates:
        if d > today:
            out[d] = "Upcoming"

    return out


def _hours_for(entry: dict) -> float:
    """Hours for one entry, derived from its own in/out times."""
    ci, co = entry.get("checkIn"), entry.get("checkOut")
    if not ci or not co:
        return 0.0
    try:
        return float(
            classify_on_checkout(
                parse_wallclock_to_ist_naive(ci),
                parse_wallclock_to_ist_naive(co),
            )["hoursWorked"]
        )
    except (TypeError, ValueError, KeyError):
        return 0.0


async def _draft_entries(
    user_id: str, dates: list[str], exempt: dict[str, str]
) -> list[dict]:
    """Prefill the week from attendance.

    Carries the employee's own check-in/out AND any work notes they already
    wrote — the previous version hardcoded empty notes, which made people
    retype work they'd already logged.
    """
    attendance: dict[str, dict] = {}
    async for r in db.attendance.find(
        {"userId": user_id, "date": {"$in": dates}}
    ):
        attendance[r["date"]] = r

    entries = []
    for d in dates:
        rec = attendance.get(d) or {}
        notes = rec.get("workNotes") or ""
        entries.append(
            {
                "date": d,
                "checkIn": iso_naive(rec.get("checkIn")),
                "checkOut": iso_naive(rec.get("checkOut")),
                "hours": float(rec.get("hoursWorked", 0.0) or 0.0),
                "attendanceType": rec.get("attendanceType"),
                "projectId": None,
                "notes": notes,
                "billable": False,
                # Context for the UI, not persisted as truth:
                "attendanceStatus": rec.get("status"),
                "exempt": d in exempt,
                "exemptReason": exempt.get(d),
                "future": d > today_ist_str(),
                "hasRecord": bool(rec),
                # True when HR/a manager filled this day in and left the work
                # notes blank — the exact case this feature exists to fix.
                "needsNotes": bool(rec) and not notes and d not in exempt,
            }
        )
    return entries


def _incomplete(entries: list[dict], exempt: dict[str, str]) -> list[str]:
    """Working days still missing an in-time, out-time or notes.

    A calendar-exempt day only stays exempt while it's genuinely empty. The
    moment an entry carries a time or a note — someone worked the Saturday,
    or came in on a holiday — it's held to the same standard as any other
    working day. Otherwise a half-filled weekend row could sail through with
    a check-in and no explanation.
    """
    today = today_ist_str()
    bad = []
    for e in entries:
        d = e.get("date")
        # A day that hasn't happened is never incomplete. Data on it is
        # rejected outright at submit, not quietly demanded here.
        if d and d > today:
            continue
        has_work = bool(
            e.get("checkIn") or e.get("checkOut") or (e.get("notes") or "").strip()
        )
        if d in exempt and not has_work:
            continue
        if (e.get("attendanceType") or "").upper() in NON_WORKING_TYPES:
            continue
        if (
            not e.get("checkIn")
            or not e.get("checkOut")
            or not (e.get("notes") or "").strip()
        ):
            bad.append(d)
    return bad


def _serialize(t: dict) -> dict:
    return {
        "id": str(t["_id"]),
        "userId": t.get("userId"),
        "weekStart": t.get("weekStart"),
        "entries": t.get("entries", []),
        "totalHours": t.get("totalHours", 0.0),
        "note": t.get("note"),
        "status": t.get("status"),
        "decidedBy": t.get("decidedBy"),
        "decisionNote": t.get("decisionNote"),
        "decidedAt": iso_naive(t.get("decidedAt")),
        "createdAt": iso_naive(t.get("createdAt")),
        "appliedAt": iso_naive(t.get("appliedAt")),
        "daysApplied": t.get("daysApplied"),
        "daysCreated": t.get("daysCreated"),
    }


# ================= USER: WEEKLY VIEW =================
@user_router.get("/my")
async def my_week(
    weekStart: str = Query(...),
    user_id: str = Depends(get_current_user),
):
    """The user's timesheet for a week.

    Returns the submitted timesheet if there is one, otherwise a draft built
    from attendance so the employee only has to fill the gaps.
    """
    dates = _week_dates(weekStart)
    exempt = await _exempt_dates(user_id, dates)

    existing = await db.timesheets.find_one({
        "userId": user_id, "weekStart": weekStart,
    })
    if existing:
        return {
            **_serialize(existing),
            "draft": False,
            "exempt": exempt,
        }

    entries = await _draft_entries(user_id, dates, exempt)
    total = round(sum(e["hours"] for e in entries), 2)
    return {
        "id": None,
        "userId": user_id,
        "weekStart": weekStart,
        "entries": entries,
        "totalHours": total,
        "status": "DRAFT",
        "draft": True,
        "exempt": exempt,
        "incompleteDates": _incomplete(entries, exempt),
    }


@user_router.post("/submit")
async def submit_timesheet(
    data: TimesheetSubmit,
    user_id: str = Depends(get_current_user),
):
    dates = _week_dates(data.weekStart)
    exempt = await _exempt_dates(user_id, dates)

    existing = await db.timesheets.find_one({
        "userId": user_id, "weekStart": data.weekStart,
    })
    # DRAFT covers a week withdrawn via /my/recall; REJECTED covers one the
    # manager sent back. Only PENDING and APPROVED are genuinely locked —
    # omitting DRAFT here meant a withdrawn sheet could never be resubmitted.
    if existing and existing.get("status") not in ("DRAFT", "REJECTED"):
        raise HTTPException(
            400,
            f"This week is {existing.get('status')}. "
            + (
                "Withdraw it first if you need to change it."
                if existing.get("status") == "PENDING"
                else "It's already approved — raise an attendance correction."
            ),
        )

    # Only the reporting manager may approve, so a week with no manager
    # assigned would sit PENDING forever. Fail now, with something the
    # employee can act on, rather than silently parking it.
    approver_id = await _reporting_manager_id(user_id)
    if not approver_id:
        raise HTTPException(
            400,
            "You don't have a reporting manager assigned, so there's nobody "
            "to approve this timesheet. Ask HR to set your reporting manager.",
        )

    if data.entries:
        for e in data.entries:
            if e.date not in dates:
                raise HTTPException(
                    400, f"entry date {e.date} not in this week",
                )
        entries = [
            {
                "date": e.date,
                "checkIn": e.checkIn,
                "checkOut": e.checkOut,
                "attendanceType": (e.attendanceType or "OFFICE").upper(),
                "projectId": e.projectId,
                "notes": (e.notes or "").strip(),
                "billable": bool(e.billable),
            }
            for e in data.entries
        ]
        # Any day of the week the client didn't send at all still has to be
        # accounted for, or an incomplete day could be hidden by omission.
        sent = {e["date"] for e in entries}
        for d in dates:
            if d not in sent:
                entries.append(
                    {
                        "date": d,
                        "checkIn": None,
                        "checkOut": None,
                        "attendanceType": None,
                        "projectId": None,
                        "notes": "",
                        "billable": False,
                    }
                )
        entries.sort(key=lambda e: e["date"])
    else:
        entries = await _draft_entries(user_id, dates, exempt)

    # You cannot log work for a day that hasn't happened. Rejected loudly
    # rather than silently dropped, so nobody thinks it was recorded.
    today = today_ist_str()
    ahead = [
        e["date"]
        for e in entries
        if e["date"] > today
        and (
            e.get("checkIn")
            or e.get("checkOut")
            or (e.get("notes") or "").strip()
        )
    ]
    if ahead:
        raise HTTPException(
            400,
            "You can't log time for a day that hasn't happened yet: "
            + ", ".join(sorted(ahead)),
        )

    # An out-time that isn't after the in-time. The completeness check passes
    # (all three fields are filled), and _hours_for quietly returns 0 — so
    # without this the day reached attendance as a HALF_DAY worth 0 hours.
    reversed_days = sorted({
        e["date"]
        for e in entries
        if e.get("checkIn") and e.get("checkOut")
        and _hours_for(e) <= 0
    })
    if reversed_days:
        raise HTTPException(
            400,
            "The out-time must be after the in-time on: "
            + ", ".join(reversed_days),
        )

    # "Every field — in time, out time and notes" on every working day.
    missing = _incomplete(entries, exempt)
    if missing:
        raise HTTPException(
            400,
            "Complete the in-time, out-time and work notes for: "
            + ", ".join(missing),
        )

    # Derive hours from the times rather than trusting whatever was sent.
    total = 0.0
    for e in entries:
        e["hours"] = _hours_for(e)
        total += e["hours"]
    total = round(total, 2)

    now = now_ist_naive()
    doc = {
        "userId": user_id,
        "weekStart": data.weekStart,
        "entries": entries,
        "totalHours": total,
        "note": data.note or "",
        "status": "PENDING",
        "approverId": approver_id,
        "decidedBy": None,
        "decisionNote": "",
        "decidedAt": None,
        "appliedAt": None,
        "createdAt": (existing or {}).get("createdAt") or now,
        "updatedAt": now,
    }

    if existing:
        await db.timesheets.update_one(
            {"_id": existing["_id"]}, {"$set": doc},
        )
        record_id = existing["_id"]
    else:
        result = await db.timesheets.insert_one(doc)
        record_id = result.inserted_id

    who_doc = await db.users.find_one({"_id": ObjectId(user_id)}, {"name": 1})
    who = (who_doc or {}).get("name") or "An employee"
    title = "Timesheet submitted"
    body = (
        f"{who} submitted a timesheet for week of {data.weekStart} "
        f"({total}h)"
    )
    # Goes to the reporting manager — the only person who can decide it.
    await notify_user(
        approver_id,
        "timesheet_submitted",
        title,
        body,
        {"timesheetId": str(record_id)},
    )
    await log_audit(
        actor_id=user_id,
        action="timesheet.submit",
        entity_type="timesheets",
        entity_id=str(record_id),
        metadata={"weekStart": data.weekStart, "totalHours": total},
    )

    saved = await db.timesheets.find_one({"_id": record_id})
    return _serialize(saved)


@user_router.post("/my/recall")
async def recall_my_timesheet(
    weekStart: str = Query(...),
    user_id: str = Depends(get_current_user),
):
    """Pull a PENDING week back to DRAFT so it can be corrected.

    Submitting locks the sheet — otherwise the manager could approve one set
    of hours while the employee edits another underneath them. But while it
    is still PENDING nobody has acted on it, so being locked out of your own
    mistake is just a trap. APPROVED sheets stay locked: those have already
    been written into attendance, and changing them there is what the
    attendance correction flow is for.
    """
    doc = await db.timesheets.find_one({
        "userId": user_id, "weekStart": weekStart,
    })
    if not doc:
        raise HTTPException(404, "No timesheet for that week.")

    status = doc.get("status")
    if status == "APPROVED":
        raise HTTPException(
            400,
            "This week has already been approved and written into your "
            "attendance. Raise an attendance correction instead.",
        )
    if status != "PENDING":
        raise HTTPException(400, f"Nothing to recall — it's {status}.")

    now = now_ist_naive()
    # Only claim it if it's STILL pending: the manager may be approving it
    # in the same moment.
    claimed = await db.timesheets.update_one(
        {"_id": doc["_id"], "status": "PENDING"},
        {"$set": {"status": "DRAFT", "updatedAt": now}},
    )
    if claimed.modified_count == 0:
        raise HTTPException(
            409,
            "Your manager just acted on this timesheet — reload to see it.",
        )

    approver = doc.get("approverId") or await _reporting_manager_id(user_id)
    if approver:
        who = await _user_brief(user_id)
        await notify_user(
            approver,
            "timesheet_recalled",
            "Timesheet withdrawn",
            f"{(who or {}).get('name') or 'An employee'} withdrew their "
            f"timesheet for week of {weekStart} to make changes.",
            {"weekStart": weekStart},
        )

    await log_audit(
        actor_id=user_id,
        action="timesheet.recall",
        entity_type="timesheets",
        entity_id=str(doc["_id"]),
        metadata={"weekStart": weekStart},
    )

    return _serialize(await db.timesheets.find_one({"_id": doc["_id"]}))


@user_router.get("/my/export.xlsx")
async def export_my_week(
    weekStart: str = Query(...),
    user_id: str = Depends(get_current_user),
):
    """Download one week as a fillable .xlsx — the upload template."""
    dates = _week_dates(weekStart)
    exempt = await _exempt_dates(user_id, dates)

    doc = await db.timesheets.find_one({
        "userId": user_id, "weekStart": weekStart,
    })
    if doc:
        entries = doc.get("entries", [])
        # A stored sheet has no exempt flags; take them from the live map so
        # the greying still tells the truth.
        for e in entries:
            e["exempt"] = e.get("date") in exempt
        status = doc.get("status", "DRAFT")
    else:
        entries = await _draft_entries(user_id, dates, exempt)
        status = "DRAFT"

    who = await _user_brief(user_id)
    wb = _week_workbook(who, weekStart, entries, status)
    return _xlsx(wb, f"timesheet-{weekStart}.xlsx")


@user_router.post("/my/import")
async def import_my_week(
    weekStart: str = Query(...),
    file: UploadFile = File(...),
    user_id: str = Depends(get_current_user),
):
    """Read a filled-in .xlsx and hand the entries back for review.

    Deliberately saves NOTHING. The employee sees what was read, fixes
    anything wrong, and presses Send to manager — an upload that silently
    submitted on their behalf would be a trap.
    """
    dates = _week_dates(weekStart)

    raw = await file.read()
    if not raw:
        raise HTTPException(400, "That file is empty.")
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(400, "That file is too large (max 5 MB).")

    entries = _parse_week_workbook(raw, dates)

    exempt = await _exempt_dates(user_id, dates)
    today = today_ist_str()

    # Report problems instead of quietly dropping rows, so the employee can
    # see exactly what the file got wrong.
    ahead = [
        e["date"] for e in entries
        if e["date"] > today
        and (e.get("checkIn") or e.get("checkOut") or e.get("notes"))
    ]
    for e in entries:
        e["hours"] = _hours_for(e)
        e["exempt"] = e["date"] in exempt
        e["exemptReason"] = exempt.get(e["date"])
        e["future"] = e["date"] > today

    reversed_days = [
        e["date"] for e in entries
        if e.get("checkIn") and e.get("checkOut") and e["hours"] <= 0
    ]

    return {
        "weekStart": weekStart,
        "entries": entries,
        "totalHours": round(sum(e["hours"] for e in entries), 2),
        "incompleteDates": _incomplete(entries, exempt),
        "futureDates": sorted(ahead),
        "reversedDates": sorted(reversed_days),
        "fileName": file.filename or "",
    }


# ================= MANAGER: LIST + DECIDE =================
@manager_router.get("")
async def manager_list_timesheets(
    status: Optional[str] = Query("PENDING"),
    actor: dict = Depends(require_manager_or_hr),
):
    """Timesheets visible to the caller.

    HR sees everything (read-only — see the decide route); a manager sees
    only their own direct reports.
    """
    actor_id = str(actor["_id"])
    if actor.get("role") == "HR":
        report_ids = None
    else:
        report_ids = [
            str(u["_id"])
            async for u in db.users.find(
                {"reportingManagerId": actor_id}, {"_id": 1}
            )
        ]
        if not report_ids:
            return []

    query: dict = {}
    if status:
        query["status"] = status
    if report_ids is not None:
        query["userId"] = {"$in": report_ids}

    out = []
    async for t in db.timesheets.find(query).sort("weekStart", -1):
        item = _serialize(t)
        # Without this the review screen can only show a raw ObjectId, which
        # tells the approver nothing about whose week they're signing off.
        brief = await _user_brief(t.get("userId"))
        item["user"] = brief
        # HR can see every timesheet but decide none of them. Say so here so
        # the UI can hide the buttons instead of offering an action that 403s.
        item["canDecide"] = bool(
            brief
            and brief.get("reportingManagerId") == actor_id
            and t.get("status") == "PENDING"
        )
        out.append(item)
    return out


async def _apply_week_to_attendance(
    user_id: str, entries: list[dict], actor_id: str, timesheet_id: str,
) -> dict:
    """Write an approved week into attendance.

    Updates the days that exist and CREATES the ones that don't, so a week
    the employee never checked in for still ends up as real attendance.
    Status / hours / overtime / lateness are recomputed with the shared
    classifier, so these days are indistinguishable from normally-worked
    ones — except for the `fromTimesheetId` provenance stamp.
    """
    now = now_ist_naive()
    applied = 0
    created = 0

    today = today_ist_str()

    for e in entries:
        date = e.get("date")
        if not date:
            continue

        # Belt and braces: a timesheet submitted days ago and approved today
        # could still carry a date that was future when it was filed. Never
        # write attendance for a day that hasn't happened.
        if date > today:
            continue

        notes = (e.get("notes") or "").strip()
        att_type = (e.get("attendanceType") or "").upper() or None

        parsed_in = None
        parsed_out = None
        try:
            if e.get("checkIn"):
                parsed_in = parse_wallclock_to_ist_naive(e["checkIn"])
            if e.get("checkOut"):
                parsed_out = parse_wallclock_to_ist_naive(e["checkOut"])
        except (TypeError, ValueError):
            # A malformed timestamp shouldn't abort the rest of the week.
            continue

        # Nothing to say about this day (weekly off / holiday / leave, no
        # times and no notes) — leave attendance alone rather than writing
        # an empty row for it.
        if not parsed_in and not parsed_out and not notes:
            continue

        existing = await db.attendance.find_one(
            {"userId": user_id, "date": date}
        )

        set_fields: dict = {
            "userId": user_id,
            "date": date,
            "updatedAt": now,
            # Provenance: this day came from an approved timesheet, not from
            # someone standing at the door tapping check-in.
            "fromTimesheetId": timesheet_id,
            "fromTimesheetAt": now,
        }
        if notes:
            set_fields["workNotes"] = notes
            # Whatever HR left blank is now filled in.
            set_fields["notesPendingFromEmployee"] = False
        if att_type:
            set_fields["attendanceType"] = att_type
        elif not existing:
            set_fields["attendanceType"] = "OFFICE"
        if parsed_in is not None:
            set_fields["checkIn"] = parsed_in
        if parsed_out is not None:
            set_fields["checkOut"] = parsed_out

        merged_in = parsed_in or (existing or {}).get("checkIn")
        merged_out = parsed_out or (existing or {}).get("checkOut")
        if merged_in and merged_out:
            c = classify_on_checkout(merged_in, merged_out)
            set_fields["status"] = c["status"]
            set_fields["hoursWorked"] = c["hoursWorked"]
            set_fields["overtimeHours"] = c["overtimeHours"]
            set_fields["isLate"] = c["isLate"]
        elif merged_in:
            set_fields["status"] = "CHECKED_IN"

        if existing:
            await db.attendance.update_one(
                {"_id": existing["_id"]}, {"$set": set_fields},
            )
            applied += 1
        else:
            set_fields["createdAt"] = now
            set_fields.setdefault("hoursWorked", 0.0)
            await db.attendance.insert_one(set_fields)
            created += 1

    await log_audit(
        actor_id=actor_id,
        action="timesheet.apply_to_attendance",
        entity_type="attendance",
        entity_id=timesheet_id,
        metadata={"updated": applied, "created": created},
    )
    return {"applied": applied, "created": created}


@manager_router.post("/{id}/decide")
async def manager_decide_timesheet(
    id: str,
    data: TimesheetDecision,
    actor: dict = Depends(require_manager_or_hr),
):
    """Approve or reject — REPORTING MANAGER ONLY.

    HR is deliberately excluded even though HR can see every timesheet: HR is
    usually the one who entered the missing times, and an approver who signs
    off their own entries is not an approval.
    """
    try:
        oid = ObjectId(id)
    except (InvalidId, TypeError):
        raise HTTPException(400, "Invalid id")

    t = await db.timesheets.find_one({"_id": oid})
    if not t:
        raise HTTPException(404, "Timesheet not found")
    if t.get("status") != "PENDING":
        raise HTTPException(400, f"Already {t.get('status')}")

    try:
        employee = await db.users.find_one({"_id": ObjectId(t["userId"])})
    except (InvalidId, TypeError, KeyError):
        employee = None
    if not employee:
        raise HTTPException(404, "Employee no longer exists")

    actor_id = str(actor["_id"])
    if employee.get("reportingManagerId") != actor_id:
        raise HTTPException(
            403,
            "Only this employee's reporting manager can decide their "
            "timesheet.",
        )

    now = now_ist_naive()
    approve = data.action == "APPROVE"
    new_status = "APPROVED" if approve else "REJECTED"

    # Claim the transition atomically BEFORE touching attendance, so a double
    # tap can't apply the same week twice. Mirrors the corrections flow.
    claimed = await db.timesheets.update_one(
        {"_id": oid, "status": "PENDING"},
        {
            "$set": {
                "status": new_status,
                "decidedBy": actor_id,
                "decisionNote": data.note or "",
                "decidedAt": now,
                "updatedAt": now,
            }
        },
    )
    if claimed.modified_count == 0:
        raise HTTPException(
            409, "This timesheet was just decided by someone else."
        )

    result = {"applied": 0, "created": 0}
    if approve:
        result = await _apply_week_to_attendance(
            t["userId"], t.get("entries", []), actor_id, id,
        )
        await db.timesheets.update_one(
            {"_id": oid},
            {
                "$set": {
                    "appliedAt": now,
                    "daysApplied": result["applied"],
                    "daysCreated": result["created"],
                }
            },
        )

    title = (
        f"Timesheet approved (week of {t.get('weekStart')})"
        if approve
        else f"Timesheet rejected (week of {t.get('weekStart')})"
    )
    body = data.note or (
        f"{result['applied'] + result['created']} day(s) updated in your "
        "attendance."
        if approve
        else "Please correct the week and submit it again."
    )
    await notify_user(
        t["userId"],
        "timesheet_decision",
        title,
        body,
        {"timesheetId": id, "outcome": data.action},
    )
    await log_audit(
        actor_id=actor_id,
        action=f"timesheet.{data.action.lower()}",
        entity_type="timesheets",
        entity_id=id,
    )
    return {
        "message": f"Timesheet {new_status.lower()}",
        "daysApplied": result["applied"],
        "daysCreated": result["created"],
    }


# ================= ONE PERSON, ONE WEEK: XLSX ROUND-TRIP =================
# The download and the upload share one file shape, so "download it, fill it
# in, upload it back" works without anyone being told what the columns mean.
# Anything computed (Day, Hours) is a formula or is recomputed server-side, so
# a wrong value typed into the file can never become truth.

SHEET_NAME = "Timesheet"
ATT_TYPES = ["OFFICE", "WFH", "CLIENT", "LEAVE", "HOLIDAY"]

# 1-based column positions in the editable sheet.
COL_DATE, COL_DAY, COL_IN, COL_OUT, COL_HOURS, COL_TYPE, COL_NOTES = range(1, 8)
HEADER_ROW = 5

_HDR_FILL = PatternFill("solid", fgColor="10305F")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_LOCK_FILL = PatternFill("solid", fgColor="EEF1F5")
_OFF_FILL = PatternFill("solid", fgColor="F7F8FA")
_THIN = Side(style="thin", color="D5DBE3")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _as_time(iso):
    """ISO string -> a real Excel time value, not text."""
    if not iso:
        return None
    try:
        return parse_wallclock_to_ist_naive(iso).time()
    except (TypeError, ValueError):
        return None


def _week_workbook(who, week_start, entries, status):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    name = (who or {}).get("name") or ""
    code = (who or {}).get("employeeCode") or ""

    ws["A1"] = "Weekly timesheet"
    ws["A1"].font = Font(bold=True, size=14, color="10305F")
    ws["A2"] = "{}    {}".format(name, code)
    ws["A2"].font = Font(size=10, color="55606E")
    ws["A3"] = "Week starting {}    Status: {}".format(week_start, status)
    ws["A3"].font = Font(size=10, color="55606E")

    headers = [
        ("Date", 12), ("Day", 6), ("In", 9), ("Out", 9),
        ("Hours", 9), ("Type", 12), ("Work notes", 62),
    ]
    for i, (title, width) in enumerate(headers, start=1):
        c = ws.cell(row=HEADER_ROW, column=i, value=title)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[HEADER_ROW].height = 20

    for offset, e in enumerate(entries):
        r = HEADER_ROW + 1 + offset
        date = e.get("date")
        try:
            d_obj = datetime.strptime(date, "%Y-%m-%d")
        except (TypeError, ValueError):
            d_obj = None

        exempt = bool(e.get("exempt"))

        ws.cell(row=r, column=COL_DATE, value=d_obj or date)
        ws.cell(row=r, column=COL_DAY,
                value=d_obj.strftime("%a") if d_obj else "")
        ws.cell(row=r, column=COL_IN, value=_as_time(e.get("checkIn")))
        ws.cell(row=r, column=COL_OUT, value=_as_time(e.get("checkOut")))

        # Live formula so Hours updates while the file is being filled in.
        ci = "{}{}".format(get_column_letter(COL_IN), r)
        co = "{}{}".format(get_column_letter(COL_OUT), r)
        ws.cell(
            row=r, column=COL_HOURS,
            value='=IF(OR({a}="",{b}=""),"",ROUND(({b}-{a})*24,2))'.format(
                a=ci, b=co),
        )
        ws.cell(row=r, column=COL_TYPE, value=e.get("attendanceType") or "")
        ws.cell(row=r, column=COL_NOTES, value=e.get("notes") or "")

        for col in range(1, 8):
            cell = ws.cell(row=r, column=col)
            cell.border = _BORDER
            if col == COL_DATE:
                cell.number_format = "yyyy-mm-dd"
                cell.alignment = Alignment(horizontal="center")
                cell.fill = _LOCK_FILL
            elif col == COL_DAY:
                cell.alignment = Alignment(horizontal="center")
                cell.fill = _LOCK_FILL
            elif col in (COL_IN, COL_OUT):
                cell.number_format = "HH:MM"
                cell.alignment = Alignment(horizontal="center")
            elif col == COL_HOURS:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center")
                cell.fill = _LOCK_FILL
            elif col == COL_NOTES:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            if exempt and col in (COL_IN, COL_OUT, COL_TYPE, COL_NOTES):
                cell.fill = _OFF_FILL

    last = HEADER_ROW + len(entries)

    # Type is a dropdown, so nobody invents a value the API will reject.
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(ATT_TYPES) + '"',
        allow_blank=True,
    )
    dv.errorTitle = "Invalid type"
    dv.error = "Pick one of: " + ", ".join(ATT_TYPES)
    ws.add_data_validation(dv)
    dv.add("{c}{a}:{c}{b}".format(
        c=get_column_letter(COL_TYPE), a=HEADER_ROW + 1, b=last))

    ws.cell(row=last + 1, column=COL_HOURS - 1, value="TOTAL").font = Font(
        bold=True)
    total = ws.cell(
        row=last + 1, column=COL_HOURS,
        value="=ROUND(SUM({c}{a}:{c}{b}),2)".format(
            c=get_column_letter(COL_HOURS), a=HEADER_ROW + 1, b=last),
    )
    total.font = Font(bold=True)
    total.number_format = "0.00"

    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

    # ---- Instructions, so the upload needs no explaining ----
    gs = wb.create_sheet("How to fill this in")
    gs.column_dimensions["A"].width = 18
    gs.column_dimensions["B"].width = 86
    gs["A1"] = "How to fill this in"
    gs["A1"].font = Font(bold=True, size=13, color="10305F")
    rows = [
        ("", ""),
        ("Column", "What to enter"),
        ("Date", "Do not change - it identifies the row when you upload."),
        ("Day", "Filled in for you."),
        ("In", "Time you started, HH:MM on a 24-hour clock (e.g. 09:30)."),
        ("Out", "Time you finished, HH:MM (e.g. 18:15)."),
        ("Hours", "Calculated for you. Do not type here - it is recomputed "
                  "from In and Out when you upload."),
        ("Type", "Pick from the dropdown: " + ", ".join(ATT_TYPES) + "."),
        ("Work notes", "What you actually did. REQUIRED on every working "
                       "day - the upload is rejected without it."),
        ("", ""),
        ("Greyed rows", "A weekly off, a public holiday, or a day that "
                        "hasn't happened yet. Leave blank unless you "
                        "worked it."),
        ("Worked a day off?", "Just fill in In, Out and Work notes on that "
                              "row. It will be counted."),
        ("Future dates", "Cannot be filled in - a day that hasn't happened "
                         "is rejected on upload."),
        ("", ""),
        ("After filling", "Save the file, then press Upload on the Timesheet "
                          "screen. Your entries are loaded in for review. "
                          "Nothing is submitted until you press Send to "
                          "manager."),
    ]
    bold_keys = ("Column", "Greyed rows", "Worked a day off?",
                 "Future dates", "After filling")
    for i, (a, b) in enumerate(rows, start=2):
        ca = gs.cell(row=i, column=1, value=a)
        ca.font = Font(bold=(a in bold_keys), size=10)
        ca.alignment = Alignment(vertical="top")
        cb = gs.cell(row=i, column=2, value=b)
        cb.alignment = Alignment(wrap_text=True, vertical="top")
        cb.font = Font(size=10)

    return wb


def _cell_to_hm(value):
    """A cell from the In/Out column -> 'HH:MM', or None.

    Accepts what Excel actually produces: a real time, a datetime, or text
    someone typed such as '9:30', '09:30:00' or '9.30'.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    text = str(value).strip().replace(".", ":")
    if not text:
        return None
    parts = text.split(":")
    try:
        h = int(parts[0])
        m = int(parts[1]) if len(parts) > 1 and parts[1] != "" else 0
    except (TypeError, ValueError):
        return None
    if not (0 <= h <= 23 and 0 <= m <= 59):
        return None
    return "{:02d}:{:02d}".format(h, m)


def _cell_to_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
    except (TypeError, ValueError):
        return None


def _parse_week_workbook(data, dates):
    """Uploaded .xlsx -> entries for this week. Raises HTTPException(400)."""
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception:
        raise HTTPException(
            400,
            "That file isn't a readable .xlsx. Download the template, fill "
            "it in, and upload that.",
        )

    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    by_date = {}
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, max_col=COL_NOTES):
        date = _cell_to_date(row[COL_DATE - 1].value)
        if not date or date not in dates:
            continue
        notes = row[COL_NOTES - 1].value
        att = row[COL_TYPE - 1].value
        by_date[date] = {
            "date": date,
            "inHm": _cell_to_hm(row[COL_IN - 1].value),
            "outHm": _cell_to_hm(row[COL_OUT - 1].value),
            "attendanceType": (str(att).strip().upper() if att else None),
            "notes": (str(notes).strip() if notes else ""),
        }

    if not by_date:
        raise HTTPException(
            400,
            "No rows for this week were found in that file. Make sure you "
            "uploaded the sheet downloaded for week starting "
            + dates[0] + ".",
        )

    bad_type = sorted({
        v["attendanceType"] for v in by_date.values()
        if v["attendanceType"] and v["attendanceType"] not in ATT_TYPES
    })
    if bad_type:
        raise HTTPException(
            400,
            "Unknown Type value(s): " + ", ".join(bad_type)
            + ". Use one of: " + ", ".join(ATT_TYPES) + ".",
        )

    out = []
    for d in dates:
        v = by_date.get(d)
        if not v:
            out.append({"date": d, "checkIn": None, "checkOut": None,
                        "attendanceType": None, "notes": ""})
            continue
        out.append({
            "date": d,
            "checkIn": (d + "T" + v["inHm"] + ":00") if v["inHm"] else None,
            "checkOut": (d + "T" + v["outHm"] + ":00") if v["outHm"] else None,
            "attendanceType": v["attendanceType"],
            "notes": v["notes"],
        })
    return out


# ================= SHARED: SCOPE, TOTALS, DOWNLOAD =================
# HR and managers get the same box — "Time Sheets", with the total hours
# worked, the sheet itself, and a download. The only difference is scope:
# HR sees the whole company, a manager sees their own reports.

DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


async def _scoped_query(
    actor: dict,
    status: Optional[str],
    user_id: Optional[str],
    week_start: Optional[str],
) -> Optional[dict]:
    """Mongo query for what this actor may see. None = nothing to show."""
    query: dict = {}
    if status:
        query["status"] = status.upper()
    if week_start:
        query["weekStart"] = week_start

    if actor.get("role") == "HR":
        if user_id:
            query["userId"] = user_id
        return query

    actor_id = str(actor["_id"])
    report_ids = [
        str(u["_id"])
        async for u in db.users.find(
            {"reportingManagerId": actor_id}, {"_id": 1}
        )
    ]
    if not report_ids:
        return None
    if user_id:
        if user_id not in report_ids:
            raise HTTPException(403, "Not one of your direct reports")
        query["userId"] = user_id
    else:
        query["userId"] = {"$in": report_ids}
    return query


async def _totals(query: Optional[dict]) -> dict:
    """Headline numbers for the Time Sheets box."""
    empty = {
        "count": 0,
        "employees": 0,
        "totalHours": 0.0,
        "byStatus": {},
        "approvedHours": 0.0,
    }
    if query is None:
        return empty

    count = 0
    hours = 0.0
    approved_hours = 0.0
    by_status: dict[str, int] = {}
    users: set[str] = set()

    async for t in db.timesheets.find(
        query, {"totalHours": 1, "status": 1, "userId": 1}
    ):
        count += 1
        h = float(t.get("totalHours", 0) or 0)
        hours += h
        st = t.get("status") or "UNKNOWN"
        by_status[st] = by_status.get(st, 0) + 1
        if st == "APPROVED":
            approved_hours += h
        if t.get("userId"):
            users.add(t["userId"])

    return {
        "count": count,
        "employees": len(users),
        "totalHours": round(hours, 2),
        "byStatus": by_status,
        # The number that actually means something: hours a manager signed off.
        "approvedHours": round(approved_hours, 2),
    }


async def _timesheet_workbook(query: Optional[dict]) -> Workbook:
    """Two sheets: one row per week, and one row per day.

    The per-day sheet is the point — a week total tells you nothing about
    whether the work was described, which is what these are approved on.
    """
    wb = Workbook()

    ws = wb.active
    ws.title = "Weeks"
    ws.append([
        "Employee", "Employee code", "Week starting", "Total hours",
        "Status", "Decided by", "Decided at", "Applied to attendance",
        "Employee note", "Decision note",
    ])

    ds = wb.create_sheet("Days")
    ds.append([
        "Employee", "Employee code", "Week starting", "Date", "Day",
        "In", "Out", "Hours", "Type", "Work notes",
    ])

    for sheet in (ws, ds):
        for cell in sheet[1]:
            cell.font = cell.font.copy(bold=True)

    if query is None:
        return wb

    # Names are looked up once per employee, not once per row.
    briefs: dict[str, dict] = {}

    async def brief(uid: Optional[str]) -> dict:
        if not uid:
            return {}
        if uid not in briefs:
            briefs[uid] = await _user_brief(uid) or {}
        return briefs[uid]

    async for t in db.timesheets.find(query).sort("weekStart", -1):
        who = await brief(t.get("userId"))
        decider = await brief(t.get("decidedBy"))
        name = who.get("name") or t.get("userId", "")
        code = who.get("employeeCode") or ""

        ws.append([
            name,
            code,
            t.get("weekStart", ""),
            float(t.get("totalHours", 0) or 0),
            t.get("status", ""),
            decider.get("name") or "",
            iso_naive(t.get("decidedAt")) or "",
            iso_naive(t.get("appliedAt")) or "",
            t.get("note", "") or "",
            t.get("decisionNote", "") or "",
        ])

        for e in t.get("entries", []):
            date = e.get("date", "")
            try:
                day = DAY_NAMES[
                    datetime.strptime(date, "%Y-%m-%d").weekday()
                ]
            except (TypeError, ValueError):
                day = ""
            ci = e.get("checkIn") or ""
            co = e.get("checkOut") or ""
            ds.append([
                name,
                code,
                t.get("weekStart", ""),
                date,
                day,
                ci[11:16] if len(ci) >= 16 else ci,
                co[11:16] if len(co) >= 16 else co,
                float(e.get("hours", 0) or 0),
                e.get("attendanceType", "") or "",
                e.get("notes", "") or "",
            ])

    # Readable column widths — an unusable download is not a download.
    for sheet, widths in (
        (ws, [26, 14, 14, 12, 12, 20, 20, 20, 40, 40]),
        (ds, [26, 14, 14, 12, 6, 8, 8, 8, 10, 60]),
    ):
        for i, w in enumerate(widths, start=1):
            sheet.column_dimensions[
                sheet.cell(row=1, column=i).column_letter
            ].width = w

    return wb


def _xlsx(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


async def _overview(query):
    """One row per employee, so an approver sees people rather than sheets.

    The old screen listed sheets with a raw ObjectId and no way to see who
    was behind on what — the answer to "who hasn't sent me anything" wasn't
    on the page at all.
    """
    if query is None:
        return []

    people = {}
    async for t in db.timesheets.find(query):
        uid = t.get("userId")
        if not uid:
            continue
        row = people.get(uid)
        if row is None:
            row = {
                "userId": uid,
                "sheets": 0,
                "totalHours": 0.0,
                "approvedHours": 0.0,
                "pending": 0,
                "approved": 0,
                "rejected": 0,
                "lastWeek": None,
                "lastStatus": None,
            }
            people[uid] = row

        hours = float(t.get("totalHours", 0) or 0)
        status = t.get("status") or ""
        row["sheets"] += 1
        row["totalHours"] += hours
        if status == "APPROVED":
            row["approved"] += 1
            row["approvedHours"] += hours
        elif status == "PENDING":
            row["pending"] += 1
        elif status == "REJECTED":
            row["rejected"] += 1

        week = t.get("weekStart")
        if week and (row["lastWeek"] is None or week > row["lastWeek"]):
            row["lastWeek"] = week
            row["lastStatus"] = status

    out = []
    for uid, row in people.items():
        row["user"] = await _user_brief(uid)
        row["totalHours"] = round(row["totalHours"], 2)
        row["approvedHours"] = round(row["approvedHours"], 2)
        out.append(row)

    # Whoever needs attention first: most pending, then most recent activity.
    out.sort(
        key=lambda r: (
            -r["pending"],
            (r["user"] or {}).get("name") or "",
        )
    )
    return out


@manager_router.get("/overview")
async def manager_timesheet_overview(
    status: Optional[str] = Query(None),
    weekStart: Optional[str] = Query(None),
    actor: dict = Depends(require_manager_or_hr),
):
    return await _overview(
        await _scoped_query(actor, status, None, weekStart)
    )


@hr_router.get("/overview")
async def hr_timesheet_overview(
    status: Optional[str] = Query(None),
    weekStart: Optional[str] = Query(None),
    hr: dict = Depends(require_hr),
):
    return await _overview(await _scoped_query(hr, status, None, weekStart))


@manager_router.get("/summary")
async def manager_timesheet_summary(
    status: Optional[str] = Query(None),
    userId: Optional[str] = Query(None),
    weekStart: Optional[str] = Query(None),
    actor: dict = Depends(require_manager_or_hr),
):
    return await _totals(
        await _scoped_query(actor, status, userId, weekStart)
    )


@manager_router.get("/export.xlsx")
async def manager_export_timesheets(
    status: Optional[str] = Query(None),
    userId: Optional[str] = Query(None),
    weekStart: Optional[str] = Query(None),
    actor: dict = Depends(require_manager_or_hr),
):
    wb = await _timesheet_workbook(
        await _scoped_query(actor, status, userId, weekStart)
    )
    return _xlsx(wb, f"timesheets{'-' + weekStart if weekStart else ''}.xlsx")


@hr_router.get("/summary")
async def hr_timesheet_summary(
    status: Optional[str] = Query(None),
    userId: Optional[str] = Query(None),
    weekStart: Optional[str] = Query(None),
    hr: dict = Depends(require_hr),
):
    return await _totals(
        await _scoped_query(hr, status, userId, weekStart)
    )


@hr_router.get("/export.xlsx")
async def hr_export_timesheets(
    status: Optional[str] = Query(None),
    userId: Optional[str] = Query(None),
    weekStart: Optional[str] = Query(None),
    hr: dict = Depends(require_hr),
):
    wb = await _timesheet_workbook(
        await _scoped_query(hr, status, userId, weekStart)
    )
    return _xlsx(wb, f"timesheets{'-' + weekStart if weekStart else ''}.xlsx")


# ================= HR: ORG-WIDE VIEW =================
@hr_router.get("")
async def hr_list_timesheets(
    status: Optional[str] = Query(None),
    userId: Optional[str] = Query(None),
    weekStart: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=500),
    _hr: dict = Depends(require_hr),
):
    """Read-only org-wide view. HR cannot decide — see the decide route."""
    query: dict = {}
    if status:
        query["status"] = status
    if userId:
        query["userId"] = userId
    if weekStart:
        query["weekStart"] = weekStart
    out = []
    async for t in db.timesheets.find(query).sort(
        "weekStart", -1
    ).limit(limit):
        item = _serialize(t)
        item["user"] = await _user_brief(t.get("userId"))
        # HR never decides — say so explicitly rather than letting the client
        # infer it from the role.
        item["canDecide"] = False
        out.append(item)
    return out
