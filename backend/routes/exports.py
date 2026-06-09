"""XLSX exports for HR.

Each endpoint mirrors a JSON report in routes/reports.py but streams an
.xlsx file. Frontend can offer both "view" (JSON) and "download"
(this) without duplicating query logic.
"""

import io
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse

from datetime import datetime
from typing import Optional

from openpyxl import Workbook

from database import db
from utils.dependencies import require_hr, require_hr_or_ceo

router = APIRouter()


def _xlsx_response(
    wb: Workbook, filename: str,
) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


def _write_header(ws, columns: list[str]) -> None:
    ws.append(columns)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)


def _nested(d: dict, *path, default=""):
    """Safe getter for nested profile fields (e.g. personal.address.city).
    Returns `default` if any segment is missing or not a dict."""
    cur: object = d
    for p in path:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(p)
        if cur is None:
            return default
    return cur if cur is not None else default


# ================= USERS / EMPLOYEE DETAILS =================
@router.get("/users.xlsx")
async def export_users(
    _hr: dict = Depends(require_hr_or_ceo),
):
    # Load everyone up front so manager + department ids can be resolved to
    # human-readable names in the sheet.
    users = [u async for u in db.users.find().sort("name", 1)]
    name_by_id = {str(u["_id"]): u.get("name", "") for u in users}
    dept_by_id: dict[str, str] = {}
    async for d in db.departments.find():
        dept_by_id[str(d["_id"])] = d.get("name", "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    _write_header(ws, [
        # Identity
        "Employee Code", "Name", "Legal Name", "Email", "Role",
        "Designation", "Status", "Joining Date",
        # Org
        "Department", "Reporting Manager", "Job Title", "Employee Type",
        "Work Location", "Work Address",
        # Contact / personal
        "Personal Email", "Personal Phone", "Work Phone", "Birthday",
        "Gender", "Blood Group", "Marital Status", "Place of Birth",
        # Address
        "Address Line 1", "Address Line 2", "City", "State",
        "Pin Code", "Country",
        # Education
        "Certification Level", "Field of Study",
        # Emergency contact
        "Emergency Contact", "Emergency Relationship", "Emergency Phone",
        # Statutory
        "PAN", "UAN", "PF Account", "ESI Number",
        # Bank (primary account)
        "Bank Name", "Account Number", "IFSC", "Branch", "Account Holder",
        # Contract
        "Contract Start", "Contract End", "Wage Type", "Wage",
        "Wage Duration",
    ])

    for u in users:
        # Org ids can live either top-level (legacy) or under `work`.
        dept_id = _nested(u, "work", "departmentId") or u.get("departmentId", "")
        mgr_id = (
            _nested(u, "work", "reportingManagerId")
            or u.get("reportingManagerId", "")
        )
        banks = u.get("bankAccounts") or []
        bank = banks[0] if banks and isinstance(banks[0], dict) else {}

        ws.append([
            # Identity
            u.get("employeeCode", ""),
            u.get("name", ""),
            _nested(u, "personal", "legalName"),
            u.get("email", ""),
            u.get("role", "USER"),
            u.get("tag", ""),
            u.get("status", ""),
            u.get("joiningDate", ""),
            # Org
            dept_by_id.get(str(dept_id), str(dept_id) if dept_id else ""),
            name_by_id.get(str(mgr_id), ""),
            _nested(u, "work", "jobTitle") or _nested(u, "work", "jobPosition"),
            _nested(u, "contract", "employeeType"),
            _nested(u, "work", "workLocation"),
            _nested(u, "work", "workAddress"),
            # Contact / personal
            _nested(u, "personal", "personalEmail"),
            _nested(u, "personal", "phone"),
            u.get("workPhone", ""),
            _nested(u, "personal", "birthday"),
            _nested(u, "personal", "gender"),
            _nested(u, "personal", "bloodGroup"),
            _nested(u, "personal", "maritalStatus"),
            _nested(u, "personal", "placeOfBirth"),
            # Address
            _nested(u, "personal", "address", "street1"),
            _nested(u, "personal", "address", "street2"),
            _nested(u, "personal", "address", "city"),
            _nested(u, "personal", "address", "state"),
            _nested(u, "personal", "address", "pinCode"),
            _nested(u, "personal", "address", "country"),
            # Education
            _nested(u, "personal", "education", "certificationLevel"),
            _nested(u, "personal", "education", "fieldOfStudy"),
            # Emergency contact
            _nested(u, "emergencyContact", "contactName"),
            _nested(u, "emergencyContact", "relationship"),
            _nested(u, "emergencyContact", "phone"),
            # Statutory
            _nested(u, "statutory", "pan"),
            _nested(u, "statutory", "uan"),
            _nested(u, "statutory", "pfAccountNumber"),
            _nested(u, "statutory", "esiNumber"),
            # Bank (primary account)
            bank.get("bankName", ""),
            bank.get("accountNumber", ""),
            bank.get("ifscCode", ""),
            bank.get("branch", ""),
            bank.get("accountHolderName", ""),
            # Contract
            _nested(u, "contract", "contractStartDate"),
            _nested(u, "contract", "contractEndDate"),
            _nested(u, "contract", "wageType"),
            _nested(u, "contract", "wage"),
            _nested(u, "contract", "wageDuration"),
        ])
    return _xlsx_response(wb, "employees.xlsx")


# ================= ATTENDANCE =================
@router.get("/attendance.xlsx")
async def export_attendance(
    fromDate: Optional[str] = Query(None),
    toDate: Optional[str] = Query(None),
    _hr: dict = Depends(require_hr_or_ceo),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    _write_header(ws, [
        "userId", "date", "attendanceType", "status", "isLate",
        "checkIn", "checkOut", "hoursWorked", "overtimeHours",
        "workNotes",
    ])

    query: dict = {}
    if fromDate:
        query.setdefault("date", {})["$gte"] = fromDate
    if toDate:
        query.setdefault("date", {})["$lte"] = toDate

    async for r in db.attendance.find(query).sort("date", -1):
        ws.append([
            r.get("userId", ""),
            r.get("date", ""),
            r.get("attendanceType", ""),
            r.get("status", ""),
            r.get("isLate", False),
            r["checkIn"].isoformat() if r.get("checkIn") else "",
            r["checkOut"].isoformat() if r.get("checkOut") else "",
            float(r.get("hoursWorked", 0) or 0),
            float(r.get("overtimeHours", 0) or 0),
            r.get("workNotes", ""),
        ])
    return _xlsx_response(wb, "attendance.xlsx")


# ================= LEAVE REQUESTS =================
@router.get("/leave-requests.xlsx")
async def export_leave_requests(
    status: Optional[str] = Query(None),
    _hr: dict = Depends(require_hr_or_ceo),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leave"
    _write_header(ws, [
        "id", "userId", "leaveTypeCode", "fromDate", "toDate",
        "totalDays", "halfDay", "status", "decidedBy", "decisionNote",
    ])
    query: dict = {}
    if status:
        query["status"] = status
    async for r in db.leave_requests.find(query).sort(
        "createdAt", -1
    ):
        ws.append([
            str(r["_id"]),
            r.get("userId", ""),
            r.get("leaveTypeCode", ""),
            r.get("fromDate", ""),
            r.get("toDate", ""),
            float(r.get("totalDays", 0) or 0),
            r.get("halfDay", False),
            r.get("status", ""),
            r.get("decidedBy", ""),
            r.get("decisionNote", ""),
        ])
    return _xlsx_response(wb, "leave-requests.xlsx")


# ================= OFFICE EXPENSES =================
@router.get("/expenses.xlsx")
async def export_expenses(
    fromDate: Optional[str] = Query(None, alias="from"),
    toDate: Optional[str] = Query(None, alias="to"),
    category: Optional[str] = Query(None),
    _hr: dict = Depends(require_hr_or_ceo),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    _write_header(ws, [
        "id", "date", "title", "category", "amount",
        "vendor", "paymentMethod", "description",
        "receiptUrl", "createdAt",
    ])

    query: dict = {}
    if fromDate:
        query.setdefault("date", {})["$gte"] = fromDate
    if toDate:
        query.setdefault("date", {})["$lte"] = toDate
    if category:
        query["category"] = category

    async for e in db.expenses.find(query).sort("date", -1):
        ws.append([
            str(e["_id"]),
            e.get("date", ""),
            e.get("title", ""),
            e.get("category", ""),
            float(e.get("amount", 0) or 0),
            e.get("vendor", ""),
            e.get("paymentMethod", ""),
            e.get("description", ""),
            e.get("receiptUrl", ""),
            e["createdAt"].isoformat() if e.get("createdAt") else "",
        ])

    name = "expenses"
    if fromDate or toDate:
        name = f"expenses_{fromDate or 'start'}_{toDate or 'end'}"
    return _xlsx_response(wb, f"{name}.xlsx")


# ================= PAYROLL =================
@router.get("/payroll/{year}/{month}.xlsx")
async def export_payroll(
    year: int,
    month: int,
    _hr: dict = Depends(require_hr_or_ceo),
):
    if month < 1 or month > 12:
        raise HTTPException(400, "month must be 1..12")

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"
    _write_header(ws, [
        "userId", "name", "totalGross", "totalDeductions", "netPay", "status",
    ])
    async for p in db.payslips.find({"year": year, "month": month}):
        ws.append([
            p.get("userId", ""),
            p.get("employeeName") or p.get("name") or "",
            float(p.get("totalGross", 0) or 0),
            float(p.get("totalDeductions", 0) or 0),
            float(p.get("netPay", 0) or 0),
            p.get("status", ""),
        ])
    return _xlsx_response(wb, f"payroll-{year}-{month:02d}.xlsx")
