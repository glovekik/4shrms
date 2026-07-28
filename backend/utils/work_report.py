"""Employee work reports — clean Excel + PDF of who worked, when, and on what.

One detail row per employee per day: name, code, date, check-in / check-out,
hours, type, status, and the work-done note. Used by HR (whole company) and
managers (their direct reports) over a Daily / Weekly / Monthly range.
"""

import os
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image,
)

from config import COMPANY_NAME, COMPANY_LOGO_PATH


# Employee, Code, Date, Day, In, Out, Hours, Type, Status, Work done
COLUMNS = [
    ("Employee", 26),
    ("Emp Code", 12),
    ("Date", 12),
    ("Day", 10),
    ("Check In", 10),
    ("Check Out", 10),
    ("Hours", 8),
    ("Type", 10),
    ("Status", 13),
    ("Work Done", 52),
]

_NAVY = "10305F"


def _cell_values(r: dict) -> list:
    return [
        r.get("name") or "",
        r.get("employeeCode") or "",
        r.get("date") or "",
        r.get("day") or "",
        r.get("checkIn") or "—",
        r.get("checkOut") or "—",
        r.get("hours") if r.get("hours") is not None else "",
        r.get("type") or "",
        r.get("status") or "",
        r.get("workNotes") or "",
    ]


# ============================ EXCEL ============================

def build_work_xlsx(rows: list, title: str, subtitle: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Report"

    n = len(COLUMNS)
    hdr_fill = PatternFill("solid", fgColor=_NAVY)
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D9DDE3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    t = ws.cell(1, 1, title)
    t.font = Font(bold=True, size=14, color=_NAVY)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    s = ws.cell(2, 1, subtitle)
    s.font = Font(size=10, color="666666")

    HDR = 4
    for i, (name, width) in enumerate(COLUMNS, 1):
        c = ws.cell(HDR, i, name)
        c.fill = hdr_fill
        c.font = hdr_font
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width

    center_cols = {3, 4, 5, 6, 7, 8, 9}
    row_i = HDR + 1
    for r in rows:
        for i, v in enumerate(_cell_values(r), 1):
            c = ws.cell(row_i, i, v)
            c.border = border
            c.alignment = Alignment(
                vertical="top",
                wrap_text=(i == 10),
                horizontal="center" if i in center_cols else "left",
            )
            if i == 7 and isinstance(v, (int, float)):
                c.number_format = "0.00"
        row_i += 1

    if not rows:
        ws.merge_cells(start_row=HDR + 1, start_column=1,
                       end_row=HDR + 1, end_column=n)
        ws.cell(HDR + 1, 1, "No attendance records in this period.").font = (
            Font(italic=True, color="999999")
        )

    ws.freeze_panes = f"A{HDR + 1}"
    ws.sheet_view.showGridLines = False

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================ PDF ============================

def build_work_pdf(rows: list, title: str, subtitle: str) -> bytes:
    base = getSampleStyleSheet()
    cell = ParagraphStyle("cell", parent=base["Normal"], fontSize=7.5,
                          leading=9.5)
    hcell = ParagraphStyle("hcell", parent=base["Normal"], fontSize=8,
                           leading=10, textColor=colors.white,
                           fontName="Helvetica-Bold")

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
        title=title,
    )
    el = []

    if COMPANY_LOGO_PATH and os.path.isfile(COMPANY_LOGO_PATH):
        try:
            logo = Image(COMPANY_LOGO_PATH, width=34 * mm, height=13.2 * mm)
            logo.hAlign = "LEFT"
            el.append(logo)
        except Exception:
            pass
    el.append(Paragraph(
        f"<b>{title}</b>",
        ParagraphStyle("t", parent=base["Title"], fontSize=15,
                       textColor=colors.HexColor("#10305F"), spaceAfter=2),
    ))
    el.append(Paragraph(
        subtitle,
        ParagraphStyle("s", parent=base["Normal"], fontSize=9,
                       textColor=colors.grey, spaceAfter=8),
    ))

    # widths (sum ~277mm usable in landscape A4)
    widths = [40, 20, 20, 15, 17, 17, 14, 16, 22, 96]
    header = [Paragraph(name, hcell) for name, _ in COLUMNS]
    data = [header]
    for r in rows:
        vals = _cell_values(r)
        vals[9] = Paragraph(str(vals[9]).replace("\n", "<br/>"), cell)
        data.append([
            Paragraph(str(v), cell) if i != 9 else v
            for i, v in enumerate(vals)
        ])

    if not rows:
        data.append([Paragraph("No attendance records in this period.", cell)]
                    + ["" for _ in range(len(COLUMNS) - 1)])

    table = Table(data, colWidths=[w * mm for w in widths], repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10305F")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9DDE3")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (2, 1), (8, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F7F8FA")]),
    ]
    table.setStyle(TableStyle(style))
    el.append(table)

    doc.build(el)
    return buf.getvalue()
